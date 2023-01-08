import openpyxl
import json

def ytdlp_json_read_comments(json_path):
    with open(json_path, 'r', encoding='utf8') as f:
        json_data = json.load(f)

    json_comments = json_data["comments"]
    return json_comments

def ytdlp_json_to_excel(json_path, xslx_path):

    json_comments = ytdlp_json_read_comments(json_path)

    excel_wb = openpyxl.Workbook()

    worksheet = excel_wb.active

    for i, json_comment in enumerate(json_comments):        
        row_index = i + 2

        if i == 0:
            keys = json_comment.keys()
            header_column_index = 1
            for key in keys:
                header_cell = worksheet.cell(row=i + 1, column=header_column_index)
                header_cell.value = key
                header_cell.font = openpyxl.styles.Font(bold=True)
                header_column_index += 1
        
        column_index = 1
        values = json_comment.values()
        for value in values:
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.value = value
            column_index += 1       

    excel_wb.save(xslx_path)
    

def json_comments_group_by_author(json_comments):
    author_table = {}

    for i, json_comment in enumerate(json_comments):
        author_id = json_comment["author_id"]
        if author_id not in author_table:
            author_table[author_id] = [json_comment]
        else:
            author_table[author_id].append(json_comment)

    return author_table

def ytdlp_json_to_excel_group_by_author(json_path, xslx_path):
    json_comments = ytdlp_json_read_comments(json_path)

    author_table = json_comments_group_by_author(json_comments)

    excel_wb = openpyxl.Workbook()

    worksheet = excel_wb.active

    # Writing table header
    worksheet.cell(row=1, column=1).value = "author_id"
    worksheet.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)

    worksheet.cell(row=1, column=2).value = "author"
    worksheet.cell(row=1, column=2).font = openpyxl.styles.Font(bold=True)

    worksheet.cell(row=1, column=3).value = "all_texts"
    worksheet.cell(row=1, column=3).font = openpyxl.styles.Font(bold=True)

    worksheet.cell(row=1, column=4).value = "all_likes"
    worksheet.cell(row=1, column=4).font = openpyxl.styles.Font(bold=True)

    worksheet.cell(row=1, column=5).value = "max_likes"
    worksheet.cell(row=1, column=5).font = openpyxl.styles.Font(bold=True)

    row_index = 2
    for author_id, comments in author_table.items():
        worksheet.cell(row=row_index, column=1).value = author_id
        
        if comments:
            worksheet.cell(row=row_index, column=2).value = comments[0]["author"]

        all_texts = ""
        all_likes = 0
        max_likes = 0
        for comment in comments:
            all_texts += comment["text"] + "\n----\n"
            all_likes += int(comment["like_count"])            
            if max_likes < int(comment["like_count"]):
                max_likes = int(comment["like_count"])

        worksheet.cell(row=row_index, column=3).value = all_texts
        worksheet.cell(row=row_index, column=4).value = all_likes
        worksheet.cell(row=row_index, column=5).value = max_likes

        row_index += 1

    excel_wb.save(xslx_path)
